from flask import Blueprint, render_template, redirect, url_for, session, request, flash, current_app, jsonify
from functools import wraps
from werkzeug.utils import secure_filename
import os
from datetime import datetime
from sqlalchemy.orm import joinedload
from models import db, User, Resident, Announcement, CertificateRequest, Complaint, BarangayOfficial, ActivityLog
from dotenv import load_dotenv

load_dotenv()

admin_bp = Blueprint("admin_bp", __name__, template_folder="../templates/admin")

# --- Decorator for admin-only routes ---
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get("admin_logged_in"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated_function

# === Admin Pages ===
@admin_bp.route("/admindashboard")
@admin_required
def admindashboard():
    try:
        # Get actual counts from database
        announcements_count = Announcement.query.count()
        residents_count = Resident.query.count()
        officials_count = BarangayOfficial.query.count()
        certificates_count = CertificateRequest.query.count()
        complaints_count = Complaint.query.count()
    except Exception as e:
        print("Error fetching dashboard counts:", e)
        announcements_count = 0
        residents_count = 0
        officials_count = 0
        certificates_count = 0
        complaints_count = 0

    return render_template("admin/admin_dashboard.html",
                         title="Dashboard",
                         announcements_count=announcements_count,
                         residents_count=residents_count,
                         officials_count=officials_count,
                         certificates_count=certificates_count,
                         complaints_count=complaints_count)

@admin_bp.route("/adminresidents")
@admin_required
def adminresidents():
    try:
        residents = Resident.query.order_by(Resident.created_at.desc()).all()
    except Exception as e:
        print("Error fetching residents:", e)
        residents = []

    return render_template("admin_resident_management.html", title="Residents Management", residents=residents)

@admin_bp.route('/add_resident', methods=['GET', 'POST'])
@admin_required
def add_resident():
    if request.method == 'POST':
        try:
            # Get form data
            resident_id = request.form.get('resident_id')
            name = request.form.get('name')
            age = request.form.get('age')
            address = request.form.get('address')
            purok = request.form.get('purok')
            contact = request.form.get('contact')
            gender = request.form.get('gender')
            civil_status = request.form.get('civil_status')
            occupation = request.form.get('occupation')

            # Create new resident using SQLAlchemy
            new_resident = Resident(
                resident_id=resident_id,
                name=name,
                age=int(age),
                purok=purok,
                contact=contact,
                gender=gender,
                civil_status=civil_status,
                occupation=occupation
            )

            db.session.add(new_resident)
            db.session.commit()

            # Log activity
            activity_log = ActivityLog(
                admin_id=session['user_id'],
                action="Added resident",
                details=f"Added resident {name} (ID: {resident_id})"
            )
            db.session.add(activity_log)
            db.session.commit()

            flash(f"Resident {name} added successfully!", "success")
            return redirect(url_for('admin_bp.adminresidents'))

        except Exception as e:
            db.session.rollback()
            flash(f"Error adding resident: {str(e)}", "danger")
            print("Add resident error:", e)

    return render_template('admin_add_resident.html')

@admin_bp.route('/edit_resident/<int:resident_id>', methods=['GET', 'POST'])
@admin_required
def edit_resident(resident_id):
    if request.method == 'POST':
        try:
            # Get form data
            resident_id_form = request.form.get('resident_id')
            name = request.form.get('name')
            age = request.form.get('age')
            address = request.form.get('address')
            purok = request.form.get('purok')
            contact = request.form.get('contact')
            gender = request.form.get('gender')
            civil_status = request.form.get('civil_status')
            occupation = request.form.get('occupation')

            # Get existing resident
            existing_resident = Resident.query.get(resident_id)
            if not existing_resident:
                flash("Resident not found.", "danger")
                return redirect(url_for('admin_bp.adminresidents'))

            # Update resident data
            existing_resident.resident_id = resident_id_form
            existing_resident.name = name
            existing_resident.age = int(age)
            existing_resident.purok = purok
            existing_resident.contact = contact
            existing_resident.gender = gender
            existing_resident.civil_status = civil_status
            existing_resident.occupation = occupation

            db.session.commit()

            # Log activity
            activity_log = ActivityLog(
                admin_id=session['user_id'],
                action="Updated resident",
                details=f"Updated resident {name} (ID: {resident_id_form})"
            )
            db.session.add(activity_log)
            db.session.commit()

            flash(f"Resident {name} updated successfully!", "success")
            return redirect(url_for('admin_bp.adminresidents'))

        except Exception as e:
            db.session.rollback()
            flash(f"Error updating resident: {str(e)}", "danger")
            print("Edit resident error:", e)

    # GET request: fetch resident data and render form
    try:
        resident = Resident.query.get(resident_id)
        if not resident:
            flash("Resident not found.", "danger")
            return redirect(url_for('admin_bp.adminresidents'))
    except Exception as e:
        flash(f"Error fetching resident: {str(e)}", "danger")
        return redirect(url_for('admin_bp.adminresidents'))

    return render_template('admin_add_resident.html', resident=resident, edit=True)

@admin_bp.route('/delete_resident/<int:resident_id>', methods=['POST'])
@admin_required
def delete_resident(resident_id):
    try:
        resident = Resident.query.get(resident_id)
        if not resident:
            flash("Resident not found.", "danger")
            return redirect(url_for('admin_bp.adminresidents'))

        # Capture details before deletion
        resident_name = resident.name
        resident_id_val = resident.resident_id

        db.session.delete(resident)
        db.session.commit()

        # Log activity
        activity_log = ActivityLog(
            admin_id=session['user_id'],
            action="Deleted resident",
            details=f"Deleted resident {resident_name} (ID: {resident_id_val})"
        )
        db.session.add(activity_log)
        db.session.commit()

        flash("Resident deleted successfully!", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Error deleting resident: {str(e)}", "danger")
        print("Delete resident error:", e)

    return redirect(url_for('admin_bp.adminresidents'))

@admin_bp.route('/add_bulk_resident', methods=['GET', 'POST'])
@admin_required
def add_bulk_resident():
    if request.method == 'POST':
        file = request.files.get('file')
        if file:
            filename = file.filename.lower()
            if filename.endswith('.csv'):
                # Process CSV file
                try:
                    import csv
                    import io

                    # Read CSV content
                    stream = io.StringIO(file.stream.read().decode("UTF8"), newline=None)
                    csv_input = csv.reader(stream)

                    # Skip header row
                    next(csv_input, None)

                    added_count = 0
                    errors = []

                    for row in csv_input:
                        try:
                            if len(row) >= 10:  # Ensure minimum required fields
                                resident_id, first_name, last_name, middle_initial, age, purok, contact, gender, civil_status, occupation = row[:10]

                                # Construct full name
                                full_name = f"{first_name.strip()} {middle_initial.strip()} {last_name.strip()}".strip()
                                if middle_initial.strip() == "":
                                    full_name = f"{first_name.strip()} {last_name.strip()}"

                                new_resident = Resident(
                                    resident_id=resident_id.strip(),
                                    first_name=first_name.strip(),
                                    last_name=last_name.strip(),
                                    middle_initial=middle_initial.strip() if middle_initial.strip() else None,
                                    name=full_name,
                                    age=int(age.strip()) if age.strip() else 0,
                                    purok=purok.strip(),
                                    contact=contact.strip(),
                                    gender=gender.strip(),
                                    civil_status=civil_status.strip(),
                                    occupation=occupation.strip()
                                )

                                db.session.add(new_resident)
                                added_count += 1
                            else:
                                errors.append(f"Row {csv_input.line_num}: Insufficient data")
                        except Exception as e:
                            errors.append(f"Row {csv_input.line_num}: {str(e)}")

                    db.session.commit()

                    if added_count > 0:
                        flash(f"Successfully added {added_count} residents!", "success")
                    if errors:
                        flash(f"Errors encountered: {', '.join(errors[:5])}", "warning")  # Show first 5 errors

                    return redirect(url_for('admin_bp.adminresidents'))

                except Exception as e:
                    db.session.rollback()
                    flash(f"Error processing CSV file: {str(e)}", "danger")
                    print("Bulk add resident error:", e)

            elif filename.endswith('.xlsx'):
                # Process Excel file
                try:
                    from openpyxl import load_workbook
                    import io

                    # Load workbook from file stream
                    wb = load_workbook(filename=io.BytesIO(file.read()))
                    ws = wb.active

                    # Skip header row (assuming row 1 is headers)
                    rows = list(ws.iter_rows(values_only=True))[1:]

                    added_count = 0
                    errors = []

                    for row_idx, row in enumerate(rows, start=2):  # Start from row 2 (after header)
                        try:
                            if len(row) >= 10:  # Ensure minimum required fields
                                resident_id, first_name, last_name, middle_initial, age, purok, contact, gender, civil_status, occupation = row[:10]

                                # Convert to string and handle None values
                                resident_id = str(resident_id).strip() if resident_id else ""
                                first_name = str(first_name).strip() if first_name else ""
                                last_name = str(last_name).strip() if last_name else ""
                                middle_initial = str(middle_initial).strip() if middle_initial else ""
                                age = str(age).strip() if age else "0"
                                purok = str(purok).strip() if purok else ""
                                contact = str(contact).strip() if contact else ""
                                gender = str(gender).strip() if gender else ""
                                civil_status = str(civil_status).strip() if civil_status else ""
                                occupation = str(occupation).strip() if occupation else ""

                                # Construct full name
                                full_name = f"{first_name} {middle_initial} {last_name}".strip()
                                if middle_initial == "":
                                    full_name = f"{first_name} {last_name}"

                                new_resident = Resident(
                                    resident_id=resident_id,
                                    first_name=first_name,
                                    last_name=last_name,
                                    middle_initial=middle_initial if middle_initial else None,
                                    name=full_name,
                                    age=int(age) if age else 0,
                                    purok=purok,
                                    contact=contact,
                                    gender=gender,
                                    civil_status=civil_status,
                                    occupation=occupation
                                )

                                db.session.add(new_resident)
                                added_count += 1
                            else:
                                errors.append(f"Row {row_idx}: Insufficient data")
                        except Exception as e:
                            errors.append(f"Row {row_idx}: {str(e)}")

                    db.session.commit()

                    if added_count > 0:
                        flash(f"Successfully added {added_count} residents!", "success")
                    if errors:
                        flash(f"Errors encountered: {', '.join(errors[:5])}", "warning")  # Show first 5 errors

                    return redirect(url_for('admin_bp.adminresidents'))

                except Exception as e:
                    db.session.rollback()
                    flash(f"Error processing Excel file: {str(e)}", "danger")
                    print("Bulk add resident error:", e)
            else:
                flash("Please upload a valid CSV or Excel (.xlsx) file.", "danger")
        else:
            flash("No file uploaded.", "danger")

    return render_template('admin_add_bulk_resident.html')


UPLOAD_FOLDER = 'static/uploads/officials'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg'}

# Ensure upload directory exists
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@admin_bp.route("/adminofficials")
@admin_required
def adminofficials():
    try:
        officials = BarangayOfficial.query.order_by(BarangayOfficial.created_at.desc()).all()
    except Exception as e:
        print("Error fetching officials:", e)
        officials = []

    return render_template("admin_brgyofficial.html", title="Barangay Officials", officials=officials)

@admin_bp.route('/add_official', methods=['GET', 'POST'])
@admin_required
def add_official():
    if request.method == 'POST':
        full_name = request.form.get('full_name')
        position = request.form.get('position')
        contact_number = request.form.get('contact_number')
        term_start = request.form.get('term_start')
        term_end = request.form.get('term_end')
        photo = request.files.get('photo')

        photo_url = None

        # Handle photo upload
        if photo and allowed_file(photo.filename):
            filename = secure_filename(photo.filename)
            filepath = os.path.join(UPLOAD_FOLDER, filename)
            photo.save(filepath)
            photo_url = f"/static/uploads/officials/{filename}"  # public URL for template use

        try:
            new_official = BarangayOfficial(
                full_name=full_name,
                position=position,
                contact_number=contact_number,
                term_start=datetime.strptime(term_start, '%Y-%m-%d').date() if term_start else None,
                term_end=datetime.strptime(term_end, '%Y-%m-%d').date() if term_end else None,
                photo_url=photo_url
            )

            db.session.add(new_official)
            db.session.commit()

            # Log activity
            activity_log = ActivityLog(
                admin_id=session['user_id'],
                action="Added official",
                details=f"Added official {full_name} (Position: {position})"
            )
            db.session.add(activity_log)
            db.session.commit()

            flash(f"Official {full_name} added successfully!", "success")
            return redirect(url_for("admin_bp.adminofficials"))

        except Exception as e:
            db.session.rollback()
            flash(f"Error adding official: {str(e)}", "danger")
            print("Add official error:", e)

    return render_template("add_editofficial.html")

@admin_bp.route('/edit_official/<int:official_id>', methods=['GET', 'POST'])
@admin_required
def edit_official(official_id):
    if request.method == 'POST':
        full_name = request.form.get('full_name')
        position = request.form.get('position')
        contact_number = request.form.get('contact_number')
        term_start = request.form.get('term_start')
        term_end = request.form.get('term_end')
        photo = request.files.get('photo')

        # Get existing official data
        try:
            existing_official = BarangayOfficial.query.get(official_id)
            if not existing_official:
                flash("Official not found.", "danger")
                return redirect(url_for("admin_bp.adminofficials"))
        except Exception as e:
            flash(f"Error fetching official: {str(e)}", "danger")
            return redirect(url_for("admin_bp.adminofficials"))

        photo_url = existing_official.photo_url

        # Handle photo upload (replace if new photo uploaded)
        if photo and allowed_file(photo.filename):
            filename = secure_filename(photo.filename)
            filepath = os.path.join(UPLOAD_FOLDER, filename)
            photo.save(filepath)
            photo_url = f"/static/uploads/officials/{filename}"  # public URL for template use

        try:
            existing_official.full_name = full_name
            existing_official.position = position
            existing_official.contact_number = contact_number
            existing_official.term_start = datetime.strptime(term_start, '%Y-%m-%d').date() if term_start else None
            existing_official.term_end = datetime.strptime(term_end, '%Y-%m-%d').date() if term_end else None
            existing_official.photo_url = photo_url

            db.session.commit()

            # Log activity
            activity_log = ActivityLog(
                admin_id=session['user_id'],
                action="Updated official",
                details=f"Updated official {full_name} (Position: {position})"
            )
            db.session.add(activity_log)
            db.session.commit()

            flash(f"Official {full_name} updated successfully!", "success")
            return redirect(url_for("admin_bp.adminofficials"))

        except Exception as e:
            db.session.rollback()
            flash(f"Error updating official: {str(e)}", "danger")
            print("Edit official error:", e)

    # GET request: fetch official data and render form
    try:
        official = BarangayOfficial.query.get(official_id)
        if not official:
            flash("Official not found.", "danger")
            return redirect(url_for("admin_bp.adminofficials"))
    except Exception as e:
        flash(f"Error fetching official: {str(e)}", "danger")
        return redirect(url_for("admin_bp.adminofficials"))

    return render_template("add_editofficial.html", official=official, edit=True)

@admin_bp.route('/delete_official/<int:official_id>', methods=['POST'])
@admin_required
def delete_official(official_id):
    try:
        official = BarangayOfficial.query.get(official_id)
        if not official:
            flash("Official not found.", "danger")
            return redirect(url_for("admin_bp.adminofficials"))

        db.session.delete(official)
        db.session.commit()
        flash("Official deleted successfully!", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Error deleting official: {str(e)}", "danger")
        print("Delete official error:", e)

    return redirect(url_for("admin_bp.adminofficials"))

@admin_bp.route("/admincomplaints")
@admin_required
def admincomplaints():
    try:
        complaints = Complaint.query.options(joinedload(Complaint.user)).order_by(Complaint.created_at.desc()).all()
    except Exception as e:
        print("Error fetching complaints:", e)
        complaints = []

    return render_template("admin_complaints.html", title="Complaints & Blotter", complaints=complaints)

@admin_bp.route("/complaint/<int:complaint_id>/status", methods=["POST"])
@admin_required
def update_complaint_status(complaint_id):
    """Update complaint status (in progress, resolved, dismissed)"""
    try:
        action = request.form.get("action")
        if action not in ["in_progress", "resolved", "dismissed"]:
            flash("Invalid action", "danger")
            return redirect(url_for('admin_bp.admincomplaints'))

        # Check if complaint exists
        existing = Complaint.query.get(complaint_id)
        if not existing:
            flash("Complaint not found", "danger")
            return redirect(url_for('admin_bp.admincomplaints'))

        current_status = existing.status

        # Determine new status based on action
        if action == "in_progress":
            if current_status != "pending":
                flash("Complaint must be pending to start investigation", "danger")
                return redirect(url_for('admin_bp.admincomplaints'))
            new_status = "under investigation"
        elif action == "resolved":
            if current_status != "under investigation":
                flash("Complaint must be under investigation to resolve", "danger")
                return redirect(url_for('admin_bp.admincomplaints'))
            new_status = "resolved"
        elif action == "dismissed":
            if current_status not in ["pending", "under investigation"]:
                flash("Complaint must be pending or under investigation to dismiss", "danger")
                return redirect(url_for('admin_bp.admincomplaints'))
            new_status = "dismissed"

        # Update status in database
        existing.status = new_status
        existing.updated_at = datetime.now()
        db.session.commit()

        flash(f"Complaint {action.replace('_', ' ')} successfully", "success")
        return redirect(url_for('admin_bp.admincomplaints'))

    except Exception as e:
        db.session.rollback()
        print("Error updating complaint status:", e)
        flash("An error occurred while updating the complaint status", "danger")
        return redirect(url_for('admin_bp.admincomplaints'))

@admin_bp.route("/complaint/<int:complaint_id>/details")
@admin_required
def complaint_details(complaint_id):
    """Get complaint details for modal"""
    try:
        complaint = Complaint.query.options(joinedload(Complaint.user)).get(complaint_id)
        if not complaint:
            return {"success": False, "message": "Complaint not found"}, 404

        return {
            "success": True,
            "data": {
                "id": complaint.id,
                "complainant_name": complaint.user.full_name if complaint.user else "N/A",
                "complainant_email": complaint.user.email if complaint.user else "N/A",
                "incident_type": complaint.incident_type,
                "incident_date": complaint.incident_date.isoformat() if complaint.incident_date else None,
                "location": complaint.location,
                "description": complaint.description,
                "persons_involved": complaint.persons_involved or "N/A",
                "status": complaint.status,
                "evidence_url": complaint.evidence_url,
                "created_at": complaint.created_at.isoformat() if complaint.created_at else None
            }
        }

    except Exception as e:
        print("Error fetching complaint details:", e)
        return {"success": False, "message": "Failed to fetch details"}, 500

@admin_bp.route("/admincertificates")
@admin_required
def admincertificates():
    return render_template("admin_certificates.html", title="Certificates")

@admin_bp.route("/certificate/data")
@admin_required
def admin_certificates_data():
    """API endpoint to fetch certificate requests data for DataTable"""
    try:
        certificates = CertificateRequest.query.order_by(CertificateRequest.created_at.desc()).all()
    except Exception as e:
        print("Error fetching certificates:", e)
        certificates = []

    # Format data for DataTable
    data = []
    for cert in certificates:
        data.append({
            "id": cert.id,
            "full_name": cert.full_name,
            "certificate_type": cert.certificate_type,
            "purpose": cert.purpose,
            "status": cert.status,
            "created_at": cert.created_at.isoformat() if cert.created_at else None,
            "valid_id_file_url": cert.valid_id_file_url
        })

    return {"data": data}

@admin_bp.route("/certificate/<int:cert_id>/status", methods=["POST"])
@admin_required
def update_certificate_status(cert_id):
    """Update certificate status (approve/decline)"""
    try:
        action = request.json.get("action")
        if action not in ["approved", "declined"]:
            return {"success": False, "message": "Invalid action"}, 400

        # Check if certificate exists
        existing = CertificateRequest.query.get(cert_id)
        if not existing:
            return {"success": False, "message": "Certificate not found"}, 404

        current_status = existing.status

        # Determine new status based on action
        if action == "approved":
            if current_status != "pending":
                return {"success": False, "message": "Certificate must be pending to approve"}, 400
            new_status = "processing"
        elif action == "declined":
            if current_status != "pending":
                return {"success": False, "message": "Certificate must be pending to decline"}, 400
            new_status = "declined"

        # Update status in database
        existing.status = new_status
        existing.updated_at = datetime.now()
        db.session.commit()

        return {"success": True, "message": f"Certificate {action} successfully"}

    except Exception as e:
        db.session.rollback()
        print("Error updating certificate status:", e)
        return {"success": False, "message": str(e)}, 500

@admin_bp.route("/certificate/<int:cert_id>/ready", methods=["POST"])
@admin_required
def mark_certificate_ready(cert_id):
    """Mark certificate as ready for pickup"""
    try:
        # Check if certificate exists and is in processing status
        existing = CertificateRequest.query.get(cert_id)
        if not existing:
            return {"success": False, "message": "Certificate not found"}, 404

        current_status = existing.status
        if current_status != "processing":
            return {"success": False, "message": "Certificate must be processing to mark as ready for pickup"}, 400

        # Update status to ready for pickup
        existing.status = "ready for pickup"
        existing.updated_at = datetime.now()
        db.session.commit()

        return {"success": True, "message": "Certificate marked as ready for pickup"}

    except Exception as e:
        db.session.rollback()
        print("Error marking certificate as ready:", e)
        return {"success": False, "message": str(e)}, 500

@admin_bp.route("/certificate/<int:cert_id>/details")
@admin_required
def certificate_details(cert_id):
    """Get certificate details for modal"""
    try:
        cert = CertificateRequest.query.get(cert_id)

        if not cert:
            return {"success": False, "message": "Certificate not found"}, 404

        return {
            "success": True,
            "data": {
                "id": cert.id,
                "full_name": cert.full_name,
                "contact_number": cert.contact_number,
                "certificate_type": cert.certificate_type,
                "purpose": cert.purpose,
                "status": cert.status,
                "created_at": cert.created_at.isoformat() if cert.created_at else None,
                "valid_id_file_url": cert.valid_id_file_url
            }
        }

    except Exception as e:
        print("Error fetching certificate details:", e)
        return {"success": False, "message": "Failed to fetch details"}, 500

@admin_bp.route("/adminannouncement")
@admin_required
def adminannouncement():
    try:
        announcements = Announcement.query.order_by(Announcement.created_at.desc()).all()
    except Exception as e:
        print("Error fetching announcements:", e)
        announcements = []

    return render_template("admin_announcement.html", announcements=announcements)

@admin_bp.route("/adminaddannouncement", methods=["GET", "POST"])
@admin_required
def adminaddannouncement():
    if request.method == "POST":
        title = request.form.get("title")
        description = request.form.get("description")
        priority = request.form.get("priority")
        image = request.files.get("image")

        image_url = None
        if image and image.filename != "":
            filename = secure_filename(image.filename)
            upload_path = os.path.join(current_app.root_path, "static/uploads/announcements", filename)
            image.save(upload_path)
            image_url = f"/static/uploads/announcements/{filename}"

        # Get admin's first name
        try:
            admin = User.query.get(session["user_id"])
            admin_first_name = admin.first_name if admin and admin.first_name else "Admin"
        except Exception as e:
            print("Error fetching admin name:", e)
            admin_first_name = "Admin"

        # Save to MySQL
        try:
            new_announcement = Announcement(
                title=title,
                description=description,
                priority=priority,
                image_url=image_url,
                posted_by=admin_first_name
            )
            db.session.add(new_announcement)
            db.session.commit()

            flash(f"Announcement '{title}' added successfully!", "success")
        except Exception as e:
            db.session.rollback()
            print("Error adding announcement:", e)
            flash("Failed to add announcement.", "danger")

        return redirect(url_for("admin_bp.adminannouncement"))

    return render_template("admin_addanouncement.html", title="Add Announcement")


@admin_bp.route("/delete_announcement/<int:ann_id>", methods=["POST"])
@admin_required
def delete_announcement(ann_id):
    try:
        announcement = Announcement.query.get(ann_id)
        if not announcement:
            flash("Announcement not found.", "danger")
            return redirect(url_for("admin_bp.adminannouncement"))

        db.session.delete(announcement)
        db.session.commit()
        flash("Announcement deleted successfully!", "success")
    except Exception as e:
        db.session.rollback()
        print("Error deleting announcement:", e)
        flash("Failed to delete announcement.", "danger")

    return redirect(url_for("admin_bp.adminannouncement"))



@admin_bp.route("/setup-policies")
@admin_required
def setup_policies():
    """Route to provide SQL for setting up policies"""
    policy_sql = """
-- Enable RLS on complaints table
ALTER TABLE complaints ENABLE ROW LEVEL SECURITY;

-- Drop existing policy if exists
DROP POLICY IF EXISTS "Allow all operations" ON complaints;

-- Create policy that allows all operations
CREATE POLICY "Allow all operations" ON complaints
FOR ALL USING (true) WITH CHECK (true);
"""
    flash(f"Execute this SQL in your Supabase SQL Editor:\n{policy_sql}", "info")
    return redirect(url_for('admin_bp.admincomplaints'))

@admin_bp.route("/adminsettings")
@admin_required
def adminsettings():
    return render_template("admin_settings.html", title="Settings")


